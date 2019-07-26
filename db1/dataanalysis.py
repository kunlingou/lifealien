# -*- coding: utf-8 -*-

import os
import analysis.Factory as Factory
import util.FileUtil as FileUtil
from openpyxl import Workbook
from openpyxl.styles import PatternFill

if __name__ == '__main__':
    config = Factory.initConfig()
    # 拷贝文件夹，从sql到excel
    filePath = os.getcwd()
    analysisPath = filePath + "\\" + config.get("path").get("analysis_path")
    FileUtil.copy_tree(filePath + "\\" + config.get("path").get("excel_path"),
                       analysisPath)
    for dir in os.listdir(analysisPath):
        print('正在分析模块：{0}'.format(dir))
        target = os.path.join(analysisPath, dir)
        wb = Workbook()
        ws = wb['Sheet']
        maxRow = 0
        maxCol = 0
        for fileName in os.walk(target):
            if (len(fileName[2]) > 0):
                for name in fileName[2]:
                    if (name.endswith(".xlsx")):
                        excelE = Factory.getExcelE(fileName[0], name)
                        ws1 = wb.create_sheet(excelE.connkey)
                        souWS = excelE.file
                        maxCol = souWS.max_column if (
                            souWS.max_column > maxCol) else maxCol
                        maxRow = souWS.max_row if (
                            souWS.max_row > maxRow) else maxRow
                        for row in souWS.rows:
                            for cell in row:
                                ws1[cell.coordinate] = cell.value
        tarWS1 = wb['上级']
        tarWS2 = wb['下级']
        color_error = config.get('color').get('error')
        for i in range(maxRow):
            for j in range(maxCol):
                cell = ws.cell(row=i + 1, column=j + 1)
                if tarWS1.cell(
                        row=i + 1, column=j + 1).value == tarWS2.cell(
                            row=i + 1, column=j + 1).value:
                    cell.value = 0
                else:
                    cell.value = 1
                    cell.fill = PatternFill("solid", fgColor=color_error)
        wb.save(target + ".xlsx")
